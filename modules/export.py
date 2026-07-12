#!/usr/bin/env python3
"""
export.py - Export account data to CSV and Excel formats.
"""

import os
import csv
import logging
import threading
from datetime import datetime
from typing import List, Dict, Any, Optional

logger = logging.getLogger(__name__)


class ExportManager:
    """Manages export of account data to CSV and Excel files."""

    def __init__(self, config):
        self.config = config
        self._lock = threading.Lock()

    def export_accounts(self, accounts: List[Dict[str, Any]],
                        format: str = "both") -> List[str]:
        """Export accounts to files. Returns list of generated file paths."""
        paths = []

        if format in ("csv", "both") and self.config.csv_enabled:
            paths.append(self._export_csv(accounts))

        if format in ("excel", "both") and self.config.excel_enabled:
            paths.append(self._export_excel(accounts))

        return paths

    def _export_csv(self, accounts: List[Dict[str, Any]]) -> str:
        """Export accounts to CSV file."""
        os.makedirs(self.config.export_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"farming_accounts_{timestamp}.csv"
        filepath = os.path.join(self.config.export_dir, filename)

        if not accounts:
            logger.warning("No accounts to export")
            return ""

        try:
            with self._lock:
                with open(filepath, "w", newline="", encoding="utf-8") as f:
                    writer = csv.DictWriter(f, fieldnames=["email", "password", "status", "proxy_used", "created_at", "session_token"])
                    writer.writeheader()
                    for acc in accounts:
                        writer.writerow({
                            "email": acc.get("email", ""),
                            "password": acc.get("password", ""),
                            "status": acc.get("status", ""),
                            "proxy_used": acc.get("proxy_used", ""),
                            "created_at": acc.get("created_at", ""),
                            "session_token": acc.get("session_token", ""),
                        })

            logger.info(f"📄 CSV exported: {filepath} ({len(accounts)} accounts)")
            return filepath
        except Exception as e:
            logger.error(f"CSV export failed: {e}")
            return ""

    def _export_excel(self, accounts: List[Dict[str, Any]]) -> str:
        """Export accounts to Excel file with multiple sheets."""
        os.makedirs(self.config.export_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"farming_accounts_{timestamp}.xlsx"
        filepath = os.path.join(self.config.export_dir, filename)

        if not accounts:
            logger.warning("No accounts to export")
            return ""

        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            df = pd.DataFrame(accounts)

            # Create Excel with multiple sheets
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Sheet 1: All accounts
                df.to_excel(writer, sheet_name="All Accounts", index=False)

                # Sheet 2: Verified only
                verified = df[df['status'].str.lower().isin(['verified', 'login_verified', 'verified'])]
                if not verified.empty:
                    verified.to_excel(writer, sheet_name="Verified", index=False)

                # Sheet 3: Failed
                failed = df[df['status'].str.lower().isin(['failed', 'verification_failed', 'verification_timeout'])]
                if not failed.empty:
                    failed.to_excel(writer, sheet_name="Failed", index=False)

                # Sheet 4: Summary
                summary_data = {
                    "Metric": ["Total Accounts", "Verified", "Failed", "Success Rate"],
                    "Value": [
                        len(df),
                        len(verified),
                        len(failed),
                        f"{len(verified)/len(df)*100:.1f}%" if len(df) > 0 else "0%"
                    ]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name="Summary", index=False)

                # Format styles
                wb = writer.book
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = thin_border

                    # Auto-adjust column widths
                    for col in ws.columns:
                        max_length = 0
                        column_letter = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = max(adjusted_width, 12)

            logger.info(f"📊 Excel exported: {filepath} ({len(accounts)} accounts, {len(wb.sheetnames)} sheets)")
            return filepath
        except ImportError:
            logger.warning("pandas or openpyxl not installed, skipping Excel export")
            return ""
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            return ""

    def export_single_account(self, email: str, password: str, status: str,
                              proxy: str = "", filepath: Optional[str] = None) -> str:
        """Export a single account immediately (useful for streaming export)."""
        if filepath is None:
            os.makedirs(self.config.export_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = os.path.join(self.config.export_dir, f"account_{timestamp}.csv")

        try:
            with open(filepath, "a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                # Check if header exists
                if os.path.getsize(filepath) == 0:
                    writer.writerow(["email", "password", "status", "proxy_used", "created_at"])
                writer.writerow([email, password, status, proxy, datetime.now().isoformat()])

            return filepath
        except Exception as e:
            logger.error(f"Single account export failed: {e}")
            return ""
